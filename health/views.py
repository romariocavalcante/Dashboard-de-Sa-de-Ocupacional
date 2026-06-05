import csv
import io
import json
import unicodedata
from datetime import datetime
from collections import Counter, defaultdict
from pathlib import Path
from xml.etree.ElementTree import Element, SubElement, tostring

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.http import Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from .models import DashboardEntry, DashboardField, DashboardItem, DashboardModuleSetting, EmployeeRecord, UserDashboardLayout


SUMMARY_DASHBOARD_CARDS = [
	{"key": "month_certificates", "title": "Atestados Cadastrados", "element_id": "monthCertificates"},
	{"key": "inss_active", "title": "Afastados INSS", "element_id": "inssActive"},
	{"key": "expired_exams", "title": "Exames Vencidos", "element_id": "expiredExams"},
	{"key": "alert", "title": "Alerta", "element_id": "alert"},
	{"key": "certificates_chart", "title": "Atestados por Mês (Ano Atual)", "element_id": "certificatesChart"},
	{"key": "accidents_chart", "title": "Acidentes por Tipo", "element_id": "accidentsChart"},
	{"key": "absenteeism_rate", "title": "Absenteísmo Médio", "element_id": "absenteeismRate"},
	{"key": "exams_60", "title": "Exames vencem em 60 dias", "element_id": "exams60"},
	{"key": "exams_90", "title": "Exames vencem em 90 dias", "element_id": "exams90"},
	{"key": "top_employee", "title": "Top colaborador em atestados", "element_id": "topEmployee"},
]

SUMMARY_DASHBOARD_CARD_KEYS = {card["key"] for card in SUMMARY_DASHBOARD_CARDS}
STANDARD_MONTHLY_WORK_HOURS = 176

MODULE_REGISTRY = [
	("executive", "Visão Executiva (KPIs)", "Indicadores críticos para acompanhamento da diretoria."),
	("exams", "Exames Ocupacionais", "Status dos exames e conformidade legal dos colaboradores."),
	("leave_inss", "Afastamentos e INSS", "Visão de impacto operacional e evolução de afastamentos."),
	("accidents", "Acidentes e CAT", "Monitoramento de segurança, gravidade e registro de acidentes."),
	("absenteeism_health", "Absenteísmo e Saúde dos Colaboradores", "Saúde ocupacional ampliada, absenteísmo e alertas preventivos."),
]

SUPPORTED_SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".ods", ".csv"}

TRUE_VALUES = {"1", "true", "sim", "yes", "y", "x", "on"}

IMPORT_COLUMN_ALIASES = {
	"nome": "nome",
	"funcionario": "nome",
	"nome_funcionario": "nome",
	"chapa": "chapa",
	"cargo": "funcao",
	"funcao": "funcao",
	"setor": "secao",
	"departamento": "secao",
	"secao": "secao",
	"unidade": "unidade",
	"filial": "unidade",
	"status_saude": "status_saude",
	"status": "status_saude",
	"afastamento_ativo": "afastamento_ativo",
	"data_afastamento": "data_afastamento",
	"previsao_retorno": "previsao_retorno",
	"exame_status": "exame_status",
	"status_exame": "exame_status",
	"proximo_periodico": "proximo_periodico",
	"exame_realizado_no_mes": "exame_realizado_no_mes",
	"acidentes_quantidade": "acidentes_quantidade",
	"quantidade_acidentes": "acidentes_quantidade",
	"ultimo_tipo": "ultimo_tipo",
	"tipo_acidente": "ultimo_tipo",
	"data_ultimo_acidente": "data_ultimo_acidente",
	"cid": "cid",
	"motivo": "motivo",
	"dt_inicio": "dt_inicio",
	"dt_final": "dt_final",
	"dias_afastados": "dias_afastados",
	"dias_af": "dias_afastados",
	"area": "area",
	"area_code": "area_code",
	"rotulos_de_linha": "nome",
	"contagem_de_nome": "qtd_atestados",
	"qtd_atestados": "qtd_atestados",
	"atestados_json": "atestados_json",
	"payload_json": "payload_json",
	"ativo": "ativo",
	"ordem": "ordem",
	"id": "id",
}


def calculate_absenteeism_from_atestados(atestados):
	today = datetime.today()
	month = today.month
	year = today.year
	perdidos_dias_mes = 0

	for atestado in atestados:
		date = _parse_iso_date(atestado.get("data"))
		if not date or date.month != month or date.year != year:
			continue

		perdidos_dias_mes += max(_to_number(atestado.get("dias"), int, 0), 0)

	horas_perdidas_mes = perdidos_dias_mes * 8
	taxa_mensal = round((horas_perdidas_mes / STANDARD_MONTHLY_WORK_HOURS) * 100, 1) if STANDARD_MONTHLY_WORK_HOURS else 0.0

	return {
		"taxaMensal": taxa_mensal,
		"horasPerdidasMes": horas_perdidas_mes,
	}


def normalize_employee_payload(employee_payload):
	payload = dict(employee_payload or {})
	saude = dict(payload.get("saudeOcupacional") or {})
	atestados = list(saude.get("atestados") or [])
	saude["absenteismo"] = calculate_absenteeism_from_atestados(atestados)
	payload["saudeOcupacional"] = saude
	return payload


def get_active_employee_payloads():
	records = EmployeeRecord.objects.filter(ativo=True).order_by("ordem", "id")
	result = []
	for record in records:
		payload = normalize_employee_payload(record.payload if isinstance(record.payload, dict) else {})
		payload["id"] = record.id
		result.append(payload)
	return result


def _parse_iso_date(value):
	if not value:
		return None

	try:
		return datetime.strptime(str(value), "%Y-%m-%d")
	except ValueError:
		return None


def build_dashboard_stats(employee_payloads):
	today = datetime.today()
	current_month = today.month
	current_year = today.year

	inss_active = 0
	expired_exams = 0
	exams60 = 0
	exams90 = 0
	total_certificates = 0
	total_absenteeism = 0.0
	month_certificates_series = [0] * 12
	accidents_by_type = {}
	top_employee = {"nome": "-", "total": 0}

	for employee in employee_payloads:
		saude = employee.get("saudeOcupacional", {}) or {}
		atestados = saude.get("atestados", []) or []
		afastamento = saude.get("afastamentoINSS", {}) or {}
		exames = saude.get("exames", {}) or {}
		acidentes = saude.get("acidentes", {}) or {}
		absenteismo = saude.get("absenteismo", {}) or {}

		if afastamento.get("ativo"):
			inss_active += 1

		status = exames.get("status")
		if status == "vencido":
			expired_exams += 1
		elif status == "vence60":
			exams60 += 1
		elif status == "vence90":
			exams90 += 1

		for atestado in atestados:
			date = _parse_iso_date(atestado.get("data"))
			total_certificates += 1
			if not date:
				continue

			if date.year == current_year:
				month_certificates_series[date.month - 1] += 1

		accident_type = acidentes.get("ultimoTipo")
		if accident_type and accident_type != "Sem registro":
			accidents_by_type[accident_type] = accidents_by_type.get(accident_type, 0) + 1

		total_absenteeism += float(absenteismo.get("taxaMensal") or 0)

		if len(atestados) > top_employee["total"]:
			top_employee = {"nome": employee.get("nome", "-"), "total": len(atestados)}

	average_absenteeism = round(total_absenteeism / len(employee_payloads), 1) if employee_payloads else 0.0

	if expired_exams > 0 and inss_active > 0:
		alert_text = "Atenção: há exames vencidos e colaboradores afastados pelo INSS ativos."
	elif expired_exams > 0:
		alert_text = "Atenção: existem exames periódicos vencidos aguardando regularização."
	elif inss_active > 0:
		alert_text = "Monitorar datas de retorno dos afastamentos INSS em andamento."
	else:
		alert_text = "Sem alertas críticos hoje."

	return {
		"month_certificates": total_certificates,
		"inss_active": inss_active,
		"expired_exams": expired_exams,
		"exams60": exams60,
		"exams90": exams90,
		"absenteeism_rate": average_absenteeism,
		"top_employee": top_employee,
		"alert_text": alert_text,
		"certificates_series": month_certificates_series,
		"accidents_by_type": accidents_by_type,
	}


def ensure_default_module_settings():
	for index, (module_key, _label, _description) in enumerate(MODULE_REGISTRY):
		DashboardModuleSetting.objects.get_or_create(
			module_key=module_key,
			defaults={"order": index, "enabled": True},
		)


def _to_number(value, number_type=float, default=0):
	try:
		return number_type(value)
	except (TypeError, ValueError):
		return default


def _format_percent(value):
	return f"{value:.1f}%"


def _format_top_counter(counter_obj, limit=5):
	if not counter_obj:
		return "Sem dados"

	return ", ".join([f"{label} ({count})" for label, count in counter_obj.most_common(limit)])


def _counter_rows(counter_obj, key_label, value_label, limit=10):
	if not counter_obj:
		return []

	return [
		{key_label: label, value_label: count}
		for label, count in counter_obj.most_common(limit)
	]


def _table_metric(label, rows, headers, keys, empty_message="Sem dados"):
	return {
		"label": label,
		"value": empty_message,
		"table_rows": rows,
		"table_headers": headers,
		"table_keys": keys,
	}


def _table_metric_from_counter(label, counter_obj, key_label, value_label, headers, keys, limit=10):
	rows = _counter_rows(counter_obj, key_label, value_label, limit)
	return _table_metric(label, rows, headers, keys)


def _build_dashboard_facts(employee_payloads, dashboard_stats):
	total_colaboradores = len(employee_payloads)

	status_counter = Counter()
	cid_counter = Counter()
	afastamentos_setor = Counter()
	afastamentos_unidade = Counter()
	afastamentos_por_colaborador = Counter()
	acidentes_setor = Counter()
	acidentes_tipo = Counter()
	absenteismo_setor = defaultdict(list)
	horas_perdidas_total = 0
	dias_perdidos_total = 0
	mental_health_cids = Counter()
	mental_health_away = 0
	restricoes_medicas = 0
	aptos = 0
	inaptos = 0
	total_exames = 0

	for employee in employee_payloads:
		nome = employee.get("nome", "Colaborador")
		setor = employee.get("secao") or employee.get("setor") or "Sem secao"
		unidade = employee.get("unidade") or "Sem unidade"
		saude = employee.get("saudeOcupacional", {}) or {}
		afastamento = saude.get("afastamentoINSS", {}) or {}
		exames = saude.get("exames", {}) or {}
		acidentes = saude.get("acidentes", {}) or {}
		absenteismo = saude.get("absenteismo", {}) or {}
		atestados = saude.get("atestados", []) or []

		exame_status = exames.get("status") or "nao_informado"
		status_counter[exame_status] += 1
		total_exames += 1

		if exame_status == "vencido":
			inaptos += 1
		else:
			aptos += 1

		if saude.get("status") in {"medium", "high"}:
			restricoes_medicas += 1

		if afastamento.get("ativo"):
			afastamentos_setor[setor] += 1
			afastamentos_unidade[unidade] += 1
			afastamentos_por_colaborador[nome] += 1

		acidentes_quantidade = _to_number(acidentes.get("quantidadeAno"), int, 0)
		if acidentes_quantidade > 0:
			acidentes_setor[setor] += acidentes_quantidade
			acidentes_tipo[acidentes.get("ultimoTipo") or "Não informado"] += acidentes_quantidade

		taxa_mensal = _to_number(absenteismo.get("taxaMensal"), float, 0.0)
		horas_perdidas = _to_number(absenteismo.get("horasPerdidasMes"), float, 0.0)
		absenteismo_setor[setor].append(taxa_mensal)
		horas_perdidas_total += horas_perdidas

		for atestado in atestados:
			cid = (atestado.get("cid") or "Não informado").strip()
			dias = _to_number(atestado.get("dias"), int, 0)
			dias_perdidos_total += max(dias, 0)
			cid_counter[cid] += 1

			if cid.upper().startswith("F"):
				mental_health_cids[cid] += 1
				if afastamento.get("ativo"):
					mental_health_away += 1

	absenteismo_setor_text = {
		setor: _format_percent(sum(values) / len(values)) if values else "0.0%"
		for setor, values in absenteismo_setor.items()
	}

	acidentes_total = sum(acidentes_setor.values())
	taxa_frequencia = (acidentes_total / total_colaboradores * 100) if total_colaboradores else 0.0
	taxa_gravidade = (dias_perdidos_total / acidentes_total) if acidentes_total else 0.0

	avg_tempo_afastamento = (dias_perdidos_total / dashboard_stats["month_certificates"]) if dashboard_stats["month_certificates"] else 0.0
	percent_exames_em_dia = ((aptos / total_exames) * 100) if total_exames else 0.0
	percent_afastados = ((dashboard_stats["inss_active"] / total_colaboradores) * 100) if total_colaboradores else 0.0
	taxa_acidentes_100 = (acidentes_total / total_colaboradores * 100) if total_colaboradores else 0.0

	risco_setor_score = Counter()
	for employee in employee_payloads:
		setor = employee.get("secao") or employee.get("setor") or "Sem secao"
		status = (employee.get("saudeOcupacional", {}) or {}).get("status")
		if status == "high":
			risco_setor_score[setor] += 3
		elif status == "medium":
			risco_setor_score[setor] += 2
		else:
			risco_setor_score[setor] += 1

	return {
		"total_colaboradores": total_colaboradores,
		"status_counter": status_counter,
		"cid_counter": cid_counter,
		"afastamentos_setor": afastamentos_setor,
		"afastamentos_unidade": afastamentos_unidade,
		"afastamentos_por_colaborador": afastamentos_por_colaborador,
		"acidentes_setor": acidentes_setor,
		"acidentes_tipo": acidentes_tipo,
		"acidentes_total": acidentes_total,
		"taxa_frequencia": taxa_frequencia,
		"taxa_gravidade": taxa_gravidade,
		"dias_perdidos_total": dias_perdidos_total,
		"avg_tempo_afastamento": avg_tempo_afastamento,
		"horas_perdidas_total": horas_perdidas_total,
		"absenteismo_setor_text": absenteismo_setor_text,
		"mental_health_cids": mental_health_cids,
		"mental_health_away": mental_health_away,
		"restricoes_medicas": restricoes_medicas,
		"aptos": aptos,
		"inaptos": inaptos,
		"percent_exames_em_dia": percent_exames_em_dia,
		"percent_afastados": percent_afastados,
		"taxa_acidentes_100": taxa_acidentes_100,
		"risco_setor_score": risco_setor_score,
	}


def _module_metrics_from_facts(facts, dashboard_stats):
	month_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
	tendencia_rows = [
		{"mes": month_labels[index], "valor": value}
		for index, value in enumerate(dashboard_stats["certificates_series"])
	]
	absenteismo_setor_rows = [
		{"setor": setor, "taxa": taxa}
		for setor, taxa in facts["absenteismo_setor_text"].items()
	]
	indice_apto_rows = [
		{"status": "Aptos", "quantidade": facts["aptos"]},
		{"status": "Inaptos", "quantidade": facts["inaptos"]},
	]
	comparativo_afastamento_rows = [
		{"periodo": "Mês", "quantidade": dashboard_stats["inss_active"]},
		{"periodo": "Ano", "quantidade": sum(dashboard_stats["certificates_series"])},
	]
	exames_vencer_rows = [
		{"janela": "30 dias", "quantidade": "Não informado"},
		{"janela": "60 dias", "quantidade": dashboard_stats["exams60"]},
		{"janela": "90 dias", "quantidade": dashboard_stats["exams90"]},
	]

	executive_metrics = [
		{"label": "Total de colaboradores ativos", "value": facts["total_colaboradores"]},
		{"label": "% de colaboradores com exames em dia", "value": _format_percent(facts["percent_exames_em_dia"])},
		{"label": "% de colaboradores afastados", "value": _format_percent(facts["percent_afastados"])},
		{"label": "% de absenteísmo geral", "value": _format_percent(dashboard_stats["absenteeism_rate"])},
		{"label": "Taxa de acidentes por 100 colaboradores", "value": f"{facts['taxa_acidentes_100']:.2f}"},
		{"label": "Dias perdidos no mês", "value": facts["dias_perdidos_total"]},
		_table_metric_from_counter(
			"Ranking de setores com maior risco",
			facts["risco_setor_score"],
			"setor",
			"pontuacao",
			["Setor", "Pontuação"],
			["setor", "pontuacao"],
			3,
		),
		_table_metric("Tendência dos indicadores (12 meses)", tendencia_rows, ["Mês", "Indicador"], ["mes", "valor"]),
	]

	exam_metrics = [
		{"label": "Exames admissionais realizados", "value": facts["status_counter"].get("admissional", 0)},
		{"label": "Exames demissionais realizados", "value": facts["status_counter"].get("demissional", 0)},
		{"label": "Exames de retorno ao trabalho", "value": facts["status_counter"].get("retorno_trabalho", 0)},
		{"label": "Exames de mudança de função", "value": facts["status_counter"].get("mudanca_funcao", 0)},
		{"label": "Taxa de conformidade dos exames ocupacionais", "value": _format_percent(facts["percent_exames_em_dia"])},
		{"label": "Colaboradores sem ASO válido", "value": facts["inaptos"]},
	]

	health_metrics = [
		_table_metric_from_counter(
			"Ranking dos CIDs mais frequentes",
			facts["cid_counter"],
			"cid",
			"quantidade",
			["CID", "Ocorrências"],
			["cid", "quantidade"],
			5,
		),
		_table_metric_from_counter(
			"Principais causas de afastamento",
			facts["cid_counter"],
			"cid",
			"quantidade",
			["CID", "Ocorrências"],
			["cid", "quantidade"],
			3,
		),
		{"label": "Quantidade de colaboradores com restrições médicas", "value": facts["restricoes_medicas"]},
		{"label": "Controle de doenças ocupacionais", "value": facts["mental_health_cids"].total() if hasattr(facts["mental_health_cids"], "total") else sum(facts["mental_health_cids"].values())},
		{"label": "Casos suspeitos de doenças ocupacionais", "value": facts["mental_health_away"]},
		_table_metric("Índice de aptos x inaptos nos exames", indice_apto_rows, ["Status", "Quantidade"], ["status", "quantidade"]),
	]

	leave_metrics = [
		{"label": "Dias perdidos por afastamento", "value": facts["dias_perdidos_total"]},
		{"label": "Tempo médio de afastamento", "value": f"{facts['avg_tempo_afastamento']:.1f} dias"},
		_table_metric_from_counter(
			"Afastamentos por setor",
			facts["afastamentos_setor"],
			"setor",
			"quantidade",
			["Setor", "Afastamentos"],
			["setor", "quantidade"],
			10,
		),
		_table_metric_from_counter(
			"Afastamentos por unidade/filial",
			facts["afastamentos_unidade"],
			"unidade",
			"quantidade",
			["Unidade", "Afastamentos"],
			["unidade", "quantidade"],
			10,
		),
		_table_metric_from_counter(
			"Histórico de afastamentos por colaborador",
			facts["afastamentos_por_colaborador"],
			"colaborador",
			"quantidade",
			["Colaborador", "Afastamentos"],
			["colaborador", "quantidade"],
			10,
		),
		_table_metric("Comparativo mensal e anual de afastamentos", comparativo_afastamento_rows, ["Período", "Quantidade"], ["periodo", "quantidade"]),
	]

	accident_metrics = [
		{"label": "Acidentes com afastamento", "value": facts["afastamentos_setor"].total() if hasattr(facts["afastamentos_setor"], "total") else sum(facts["afastamentos_setor"].values())},
		{"label": "Acidentes sem afastamento", "value": max(facts["acidentes_total"] - dashboard_stats["inss_active"], 0)},
		{"label": "Taxa de frequência de acidentes", "value": f"{facts['taxa_frequencia']:.2f}"},
		{"label": "Taxa de gravidade de acidentes", "value": f"{facts['taxa_gravidade']:.2f}"},
		_table_metric_from_counter(
			"Acidentes por setor",
			facts["acidentes_setor"],
			"setor",
			"quantidade",
			["Setor", "Acidentes"],
			["setor", "quantidade"],
			10,
		),
		_table_metric_from_counter(
			"Acidentes por tipo de lesão",
			facts["acidentes_tipo"],
			"lesao",
			"quantidade",
			["Tipo de lesão", "Acidentes"],
			["lesao", "quantidade"],
			10,
		),
		{"label": "Acidentes por parte do corpo atingida", "value": "Não informado"},
		{"label": "Controle de emissão de CAT", "value": "Não informado"},
		{"label": "Dias perdidos por acidentes", "value": facts["dias_perdidos_total"]},
	]

	absenteeism_health_metrics = [
		_table_metric("Absenteísmo por setor", absenteismo_setor_rows, ["Setor", "Taxa"], ["setor", "taxa"]),
		{"label": "Absenteísmo por gestor", "value": "Não informado"},
		_table_metric_from_counter(
			"Absenteísmo por unidade",
			facts["afastamentos_unidade"],
			"unidade",
			"quantidade",
			["Unidade", "Ocorrências"],
			["unidade", "quantidade"],
			10,
		),
		_table_metric("Evolução mensal do absenteísmo", tendencia_rows, ["Mês", "Indicador"], ["mes", "valor"]),
		{"label": "Custo estimado do absenteísmo", "value": f"R$ {facts['horas_perdidas_total'] * 55:.2f}"},
		{"label": "Horas perdidas por absenteísmo", "value": f"{facts['horas_perdidas_total']:.1f}"},
		{"label": "Controle de vacinas obrigatórias", "value": "Não informado"},
		{"label": "Vacinas vencidas", "value": "Não informado"},
		{"label": "Vacinas próximas do vencimento", "value": "Não informado"},
		{"label": "Cobertura vacinal da empresa", "value": "Não informado"},
		{"label": "Colaboradores avaliados ergonomicamente", "value": "Não informado"},
		{"label": "Pendências de avaliações ergonômicas", "value": "Não informado"},
		{"label": "Não conformidades ergonômicas encontradas", "value": "Não informado"},
		{"label": "Plano de ação ergonômico", "value": "Não informado"},
		{"label": "Afastamentos relacionados à saúde mental", "value": facts["mental_health_away"]},
		_table_metric_from_counter(
			"Principais CIDs relacionados à saúde mental",
			facts["mental_health_cids"],
			"cid",
			"quantidade",
			["CID", "Ocorrências"],
			["cid", "quantidade"],
			10,
		),
		_table_metric_from_counter(
			"Evolução dos afastamentos psicológicos",
			facts["mental_health_cids"],
			"cid",
			"quantidade",
			["CID", "Ocorrências"],
			["cid", "quantidade"],
			12,
		),
		{"label": "Indicadores de acompanhamento psicológico", "value": "Não informado"},
		{"label": "Custos com exames ocupacionais", "value": f"R$ {facts['total_colaboradores'] * 120:.2f}"},
		{"label": "Custos com afastamentos", "value": f"R$ {facts['dias_perdidos_total'] * 180:.2f}"},
		{"label": "Custos com acidentes", "value": f"R$ {facts['acidentes_total'] * 500:.2f}"},
		{"label": "Custos por clínica prestadora", "value": "Não informado"},
		{"label": "Custos por unidade", "value": "Não informado"},
		_table_metric("Exames vencendo em 30, 60 e 90 dias", exames_vencer_rows, ["Janela", "Quantidade"], ["janela", "quantidade"]),
		{"label": "Exames vencidos", "value": dashboard_stats["expired_exams"]},
		{"label": "Retornos ao trabalho previstos para os próximos dias", "value": dashboard_stats["inss_active"]},
		{"label": "Vacinas vencendo", "value": "Não informado"},
		{"label": "Restrições médicas sem acompanhamento", "value": facts["restricoes_medicas"]},
		{"label": "Afastamentos próximos do prazo de alta", "value": dashboard_stats["inss_active"]},
	]

	return {
		"executive": executive_metrics,
		"exams": exam_metrics + health_metrics,
		"leave_inss": leave_metrics,
		"accidents": accident_metrics,
		"absenteeism_health": absenteeism_health_metrics,
	}


def build_dashboard_modules(employee_payloads, dashboard_stats):
	ensure_default_module_settings()
	facts = _build_dashboard_facts(employee_payloads, dashboard_stats)
	module_metrics = _module_metrics_from_facts(facts, dashboard_stats)

	module_settings = {
		item.module_key: item
		for item in DashboardModuleSetting.objects.all()
	}

	modules = []
	for default_order, (key, title, description) in enumerate(MODULE_REGISTRY):
		setting = module_settings.get(key)
		if setting and not setting.enabled:
			continue

		order = setting.order if setting else default_order
		modules.append({
			"key": key,
			"title": title,
			"description": description,
			"order": order,
			"metrics": module_metrics.get(key, []),
		})

	modules.sort(key=lambda item: item["order"])
	return modules


def get_employee_payload(employee_record):
	if not isinstance(employee_record.payload, dict):
		return {}

	return normalize_employee_payload(employee_record.payload)


def build_atestados_from_post(post_data):
	data_list = post_data.getlist("atestado_data")
	cid_list = post_data.getlist("atestado_cid")
	dias_list = post_data.getlist("atestado_dias")
	atestados = []

	for index, raw_date in enumerate(data_list):
		raw_cid = cid_list[index] if index < len(cid_list) else ""
		raw_dias = dias_list[index] if index < len(dias_list) else ""
		raw_date = raw_date.strip()
		raw_cid = raw_cid.strip()
		raw_dias = raw_dias.strip()

		if not raw_date and not raw_cid and not raw_dias:
			continue

		if not raw_date or not raw_cid or not raw_dias:
			raise ValueError("Preencha data, CID e dias de todos os atestados ou remova a linha em branco.")

		if not raw_dias.isdigit():
			raise ValueError("Os dias de afastamento devem ser um número inteiro.")

		atesto = {
			"data": raw_date,
			"cid": raw_cid,
			"dias": int(raw_dias),
		}
		atestados.append(atesto)

	return atestados


def _normalize_column_name(value):
	text = str(value or "").strip().lower()
	text = "".join(
		char for char in unicodedata.normalize("NFKD", text)
		if not unicodedata.combining(char)
	)
	return text.replace(" ", "_").replace("-", "_").replace(".", "_")


def _normalize_employee_key(value):
	text = _safe_text(value)
	if not text:
		return ""
	text = "".join(
		char for char in unicodedata.normalize("NFKD", text)
		if not unicodedata.combining(char)
	)
	return text.lower().strip()


def _safe_text(value):
	if value is None:
		return ""
	text = str(value).strip()
	if text.lower() in {"nan", "nat", "none", "null"}:
		return ""
	return text


def _to_bool(value):
	if isinstance(value, bool):
		return value
	return _safe_text(value).lower() in TRUE_VALUES


def _to_optional_int(value):
	text = _safe_text(value)
	if not text:
		return None

	if text.isdigit() or (text.startswith("-") and text[1:].isdigit()):
		return int(text)

	try:
		return int(float(text))
	except (TypeError, ValueError):
		return None


def _to_iso_date(value):
	if hasattr(value, "strftime"):
		return value.strftime("%Y-%m-%d")

	text = _safe_text(value)
	if not text:
		return None

	for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
		try:
			return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
		except ValueError:
			continue

	return None


def _rows_from_matrix(matrix_rows):
	if not matrix_rows:
		return []

	headers = [str(cell).strip() if cell is not None else "" for cell in matrix_rows[0]]
	rows = []
	for data_row in matrix_rows[1:]:
		if not any(cell not in (None, "") for cell in data_row):
			continue
		row_data = {}
		for index, header in enumerate(headers):
			cell_value = data_row[index] if index < len(data_row) else ""
			if not header:
				if cell_value and "area_code" not in row_data:
					row_data["area_code"] = cell_value
				continue
			row_data[header] = cell_value
		rows.append(row_data)
	return rows


def _read_employee_spreadsheet(uploaded_file):
	suffix = Path(uploaded_file.name).suffix.lower()
	if suffix not in SUPPORTED_SPREADSHEET_EXTENSIONS:
		raise ValueError("Formato não suportado. Envie .xlsx, .xls, .ods ou .csv.")

	uploaded_file.seek(0)

	if suffix == ".csv":
		content = uploaded_file.read()
		decoded = content.decode("utf-8-sig")
		reader = csv.DictReader(io.StringIO(decoded))
		return [dict(row) for row in reader]

	if suffix == ".xlsx":
		try:
			from openpyxl import load_workbook
		except ImportError as exc:
			raise RuntimeError("Dependência ausente: openpyxl.") from exc

		workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
		sheet_map = { _normalize_column_name(name): name for name in workbook.sheetnames }
		rows = []

		primary_sheet_name = sheet_map.get(_normalize_column_name("PLANILHA"))
		if primary_sheet_name:
			sheet = workbook[primary_sheet_name]
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))
		else:
			sheet = workbook.active
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))

		pivot_sheet_name = sheet_map.get(_normalize_column_name("QNT ATESTADO POR FUNCIONARIO"))
		if pivot_sheet_name:
			sheet = workbook[pivot_sheet_name]
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))

		return rows

	if suffix == ".xls":
		try:
			import xlrd
		except ImportError as exc:
			raise RuntimeError("Dependência ausente: xlrd.") from exc

		content = uploaded_file.read()
		workbook = xlrd.open_workbook(file_contents=content)
		sheet = workbook.sheet_by_index(0)
		matrix_rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
		return _rows_from_matrix(matrix_rows)

	try:
		from odf import teletype
		from odf.opendocument import load
		from odf.table import Table, TableCell, TableRow
	except ImportError as exc:
		raise RuntimeError("Dependência ausente: odfpy.") from exc

	content = uploaded_file.read()
	doc = load(io.BytesIO(content))
	matrix_rows = []

	for table in doc.spreadsheet.getElementsByType(Table):
		for row in table.getElementsByType(TableRow):
			row_values = []
			for cell in row.getElementsByType(TableCell):
				repeat_count = int(cell.getAttribute("numbercolumnsrepeated") or 1)
				text_value = teletype.extractText(cell).strip()
				for _ in range(repeat_count):
					row_values.append(text_value)
			if row_values:
				matrix_rows.append(row_values)
		if matrix_rows:
			break

	return _rows_from_matrix(matrix_rows)


def _read_employee_spreadsheet_path(file_path):
	suffix = Path(file_path).suffix.lower()
	if suffix not in SUPPORTED_SPREADSHEET_EXTENSIONS:
		raise ValueError("Formato não suportado. Envie .xlsx, .xls, .ods ou .csv.")

	if suffix == ".csv":
		content = Path(file_path).read_bytes()
		decoded = content.decode("utf-8-sig")
		reader = csv.DictReader(io.StringIO(decoded))
		return [dict(row) for row in reader]

	if suffix == ".xlsx":
		from openpyxl import load_workbook
		workbook = load_workbook(file_path, read_only=True, data_only=True)
		sheet_map = { _normalize_column_name(name): name for name in workbook.sheetnames }
		rows = []

		primary_sheet_name = sheet_map.get(_normalize_column_name("PLANILHA"))
		if primary_sheet_name:
			sheet = workbook[primary_sheet_name]
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))
		else:
			sheet = workbook.active
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))

		pivot_sheet_name = sheet_map.get(_normalize_column_name("QNT ATESTADO POR FUNCIONARIO"))
		if pivot_sheet_name:
			sheet = workbook[pivot_sheet_name]
			matrix_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
			rows.extend(_rows_from_matrix(matrix_rows))

		return rows

	if suffix == ".xls":
		import xlrd
		workbook = xlrd.open_workbook(filename=str(file_path))
		sheet = workbook.sheet_by_index(0)
		matrix_rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
		return _rows_from_matrix(matrix_rows)

	from odf import teletype
	from odf.opendocument import load
	from odf.table import Table, TableCell, TableRow

	doc = load(str(file_path))
	matrix_rows = []
	for table in doc.spreadsheet.getElementsByType(Table):
		for row in table.getElementsByType(TableRow):
			row_values = []
			for cell in row.getElementsByType(TableCell):
				repeat_count = int(cell.getAttribute("numbercolumnsrepeated") or 1)
				text_value = teletype.extractText(cell).strip()
				for _ in range(repeat_count):
					row_values.append(text_value)
			if row_values:
				matrix_rows.append(row_values)
		if matrix_rows:
			break

	return _rows_from_matrix(matrix_rows)


def _canonicalize_row(raw_row):
	parsed = {}
	for key, value in raw_row.items():
		normalized_key = _normalize_column_name(key)
		canonical_key = IMPORT_COLUMN_ALIASES.get(normalized_key)
		if canonical_key:
			parsed[canonical_key] = value
	return parsed


def _parse_atestados_json(raw_value):
	raw_text = _safe_text(raw_value)
	if not raw_text:
		return []

	try:
		parsed = json.loads(raw_text)
	except json.JSONDecodeError as exc:
		raise ValueError("Valor inválido em 'atestados_json'. Use uma lista JSON válida.") from exc

	if not isinstance(parsed, list):
		raise ValueError("A coluna 'atestados_json' precisa conter uma lista JSON.")

	result = []
	for atestado in parsed:
		if not isinstance(atestado, dict):
			continue

		dias = _to_optional_int(atestado.get("dias"))
		result.append({
			"data": _to_iso_date(atestado.get("data")),
			"cid": _safe_text(atestado.get("cid")),
			"dias": dias if dias is not None else 0,
			"motivo": _safe_text(atestado.get("motivo")),
			"area": _safe_text(atestado.get("area")),
		})

	return result


def _build_employee_payload_from_row(raw_row):
	row = _canonicalize_row(raw_row)

	payload_json_text = _safe_text(row.get("payload_json"))
	if payload_json_text:
		try:
			parsed_payload = json.loads(payload_json_text)
		except json.JSONDecodeError as exc:
			raise ValueError("Valor inválido em 'payload_json'.") from exc

		if not isinstance(parsed_payload, dict):
			raise ValueError("A coluna 'payload_json' precisa conter um objeto JSON.")

		return normalize_employee_payload(parsed_payload), row

	atestados = _parse_atestados_json(row.get("atestados_json"))
	dt_inicio = row.get("dt_inicio")
	dt_final = row.get("dt_final")
	dias_afastados = _to_optional_int(row.get("dias_afastados"))
	row_cid = _safe_text(row.get("cid"))
	row_qtd_atestados = _to_optional_int(row.get("qtd_atestados"))
	if dt_inicio or dias_afastados is not None or row_cid:
		motivo = _safe_text(row.get("motivo"))
		area = _safe_text(row.get("area")) or _safe_text(row.get("area_code"))
		atestados.append({
			"data": _to_iso_date(dt_inicio),
			"cid": row_cid,
			"dias": dias_afastados if dias_afastados is not None else 0,
			"motivo": motivo,
			"area": area,
		})
	qtd_atestados = row_qtd_atestados if row_qtd_atestados is not None else None
	saude = {
		"status": _safe_text(row.get("status_saude")) or "low",
		"afastamentoINSS": {
			"ativo": _to_bool(row.get("afastamento_ativo")),
			"dataAfastamento": _to_iso_date(row.get("data_afastamento")),
			"previsaoRetorno": _to_iso_date(row.get("previsao_retorno")),
		},
		"exames": {
			"status": _safe_text(row.get("exame_status")) or "em_dia",
			"proximoPeriodico": _to_iso_date(row.get("proximo_periodico")),
			"realizadoNoMes": _to_bool(row.get("exame_realizado_no_mes")),
		},
		"acidentes": {
			"quantidadeAno": _to_optional_int(row.get("acidentes_quantidade")) or 0,
			"ultimoTipo": _safe_text(row.get("ultimo_tipo")) or "Sem registro",
			"dataUltimo": _to_iso_date(row.get("data_ultimo_acidente")),
		},
		"atestados": atestados,
	}
	payload = {
		"chapa": _safe_text(row.get("chapa")),
		"nome": _safe_text(row.get("nome")),
		"funcao": _safe_text(row.get("funcao")) or _safe_text(row.get("cargo")),
		"dt_inicio": _to_iso_date(dt_inicio),
		"dt_final": _to_iso_date(dt_final),
		"dias_afastados": dias_afastados if dias_afastados is not None else 0,
		"motivo": _safe_text(row.get("motivo")),
		"cid": row_cid,
		"secao": _safe_text(row.get("secao")),
		"qtd_atestados": qtd_atestados,
		"saudeOcupacional": saude,
	}
	return normalize_employee_payload(payload), row


def _xml_safe_tag(tag_name):
	normalized = _normalize_column_name(tag_name)
	return normalized if normalized else "campo"


def _append_value_to_xml(parent, key, value):
	tag_name = _xml_safe_tag(key)

	if isinstance(value, dict):
		node = SubElement(parent, tag_name)
		for child_key, child_value in value.items():
			_append_value_to_xml(node, child_key, child_value)
		return

	if isinstance(value, list):
		node = SubElement(parent, tag_name)
		for item in value:
			_append_value_to_xml(node, "item", item)
		return

	node = SubElement(parent, tag_name)
	node.text = "" if value is None else str(value)


def _build_employees_xml(records):
	root = Element("funcionarios")
	for record in records:
		emp = SubElement(root, "funcionario")
		emp.set("id", str(record.id))
		emp.set("ativo", "true" if record.ativo else "false")
		emp.set("ordem", str(record.ordem))

		payload = normalize_employee_payload(record.payload if isinstance(record.payload, dict) else {})
		for key, value in payload.items():
			_append_value_to_xml(emp, key, value)

	return tostring(root, encoding="utf-8", xml_declaration=True)


def dashboard(request):
	hidden_dashboard_card_keys = set()
	if request.user.is_authenticated:
		layout, _ = UserDashboardLayout.objects.get_or_create(usuario=request.user)
		hidden_dashboard_card_keys = _get_hidden_dashboard_card_keys(layout)
	employees_payloads = get_active_employee_payloads()

	context = {
		"hidden_dashboard_card_keys": sorted(hidden_dashboard_card_keys),
		"employees_payloads": employees_payloads,
		"dashboard_stats": build_dashboard_stats(employees_payloads),
	}
	context["dashboard_modules"] = build_dashboard_modules(employees_payloads, context["dashboard_stats"])
	context["active_module_key"] = context["dashboard_modules"][0]["key"] if context["dashboard_modules"] else ""
	return render(request, 'health/index.html', context)


def _get_hidden_dashboard_card_keys(layout):
	return {card_key for card_key in layout.cards_ocultos_dashboard if card_key in SUMMARY_DASHBOARD_CARD_KEYS}


def _get_visible_dashboard_cards(hidden_dashboard_card_keys):
	return [card for card in SUMMARY_DASHBOARD_CARDS if card["key"] not in hidden_dashboard_card_keys]


def funcionarios(request):
	context = {
		"employees_payloads": get_active_employee_payloads(),
		"is_admin": user_is_admin(request.user),
	}
	return render(request, 'health/funcionarios/funcionarios.html', context)


@login_required
@require_POST
def adicionar_funcionario(request):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode adicionar funcionários.")

	nome = request.POST.get("nome", "").strip()
	if not nome:
		messages.error(request, "Informe o nome do funcionário.")
		return redirect("health:funcionarios")

	status = request.POST.get("status", "low")
	if status not in {"low", "medium", "high"}:
		status = "low"

	payload = {
		"nome": nome,
		"funcao": request.POST.get("funcao", "").strip(),
		"secao": request.POST.get("secao", "").strip(),
		"unidade": request.POST.get("unidade", "").strip(),
		"saudeOcupacional": {
			"status": status,
			"afastamentoINSS": {
				"ativo": False,
				"dataAfastamento": None,
				"previsaoRetorno": None,
			},
			"exames": {
				"status": "em_dia",
				"proximoPeriodico": None,
				"realizadoNoMes": False,
			},
			"acidentes": {
				"quantidadeAno": 0,
				"ultimoTipo": "Sem registro",
				"dataUltimo": None,
			},
			"atestados": [],
		},
	}
	payload = normalize_employee_payload(payload)
	last_order = EmployeeRecord.objects.order_by("-ordem").values_list("ordem", flat=True).first() or 0
	EmployeeRecord.objects.create(payload=payload, ativo=True, ordem=last_order + 1)
	messages.success(request, "Funcionário adicionado com sucesso.")
	return redirect("health:funcionarios")


@login_required
@require_POST
def remover_funcionario(request, employee_id):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode remover funcionários.")

	record = get_object_or_404(EmployeeRecord, id=employee_id)
	record.delete()
	messages.success(request, "Funcionário removido.")
	return redirect("health:funcionarios")


@login_required
def editar_funcionario(request, employee_id):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode editar funcionários.")

	record = get_object_or_404(EmployeeRecord, id=employee_id)
	payload = get_employee_payload(record)
	saude = payload.get("saudeOcupacional", {})
	af = saude.get("afastamentoINSS", {})
	exames = saude.get("exames", {})
	acidentes = saude.get("acidentes", {})
	absenteismo = saude.get("absenteismo", {})

	if request.method == "POST":
		try:
			updated_atestados = build_atestados_from_post(request.POST)
		except ValueError as exc:
			messages.error(request, str(exc))
		else:
			payload["nome"] = request.POST.get("nome", "").strip()
			payload["funcao"] = request.POST.get("funcao", "").strip()
			payload["secao"] = request.POST.get("secao", "").strip()
			saude["status"] = request.POST.get("status", "low")
			saude["afastamentoINSS"] = {
				"ativo": request.POST.get("afastamento_ativo") == "on",
				"dataAfastamento": request.POST.get("data_afastamento") or None,
				"previsaoRetorno": request.POST.get("previsao_retorno") or None,
			}
			saude["exames"] = {
				"status": request.POST.get("exame_status", "em_dia"),
				"proximoPeriodico": request.POST.get("proximo_periodico") or None,
				"realizadoNoMes": request.POST.get("exame_realizado_no_mes") == "on",
			}
			saude["acidentes"] = {
				"quantidadeAno": int(request.POST.get("acidentes_quantidade", "0") or 0),
				"ultimoTipo": request.POST.get("ultimo_tipo", "").strip(),
				"dataUltimo": request.POST.get("data_ultimo_acidente") or None,
			}
			saude["atestados"] = updated_atestados
			saude["absenteismo"] = calculate_absenteeism_from_atestados(updated_atestados)
			payload["saudeOcupacional"] = saude
			record.payload = payload
			record.save(update_fields=["payload"])
			messages.success(request, "Funcionário atualizado com sucesso.")
			return redirect("health:funcionarios")

	context = {
		"employee": payload,
		"employee_id": record.id,
		"status_options": [
			("low", "Estável"),
			("medium", "Atenção"),
			("high", "Crítico"),
		],
		"exame_status_options": [
			("em_dia", "Em dia"),
			("vence60", "Vence em 60 dias"),
			("vence90", "Vence em 90 dias"),
			("vencido", "Vencido"),
		],
		"atestados": saude.get("atestados", []),
		"afastamento": af,
		"exames": exames,
		"acidentes": acidentes,
		"absenteismo": absenteismo,
	}
	return render(request, "health/funcionarios/funcionario-editar.html", context)


def funcionarios_detalhes(request):
	context = {
		"employees_payloads": get_active_employee_payloads(),
		"is_admin": user_is_admin(request.user),
	}
	return render(request, 'health/funcionarios/funcionarios-detalhes.html', context)


def relatorios(request):
	context = {
		"employees_payloads": get_active_employee_payloads(),
	}
	return render(request, 'health/relatorios/relatorios.html', context)


def configuracoes(request):
	return render(request, 'health/configuracoes/configuracoes.html')


def user_is_admin(user):
	return user.is_superuser or user.is_staff or user.groups.filter(name="administrador").exists()


def ensure_default_dashboard_structure():
	if DashboardItem.objects.exists():
		return

	defaults = [
		("Atestados Médicos", ["CID", "Quantidade", "Observações"]),
		("Afastamentos INSS", ["Data de afastamento", "Previsão de retorno"]),
		("Exames Periódicos", ["Status do exame", "Data limite"]),
		("Acidentes", ["Tipo", "Descrição"]),
		("Absenteísmo", ["Taxa mensal", "Horas perdidas"]),
	]

	for item_index, (item_name, fields) in enumerate(defaults):
		item = DashboardItem.objects.create(nome=item_name, posicao=item_index)
		for field_index, field_name in enumerate(fields):
			DashboardField.objects.create(item=item, nome=field_name, posicao=field_index)


@login_required
def painel_usuario(request):
	ensure_default_dashboard_structure()

	is_admin = user_is_admin(request.user)
	layout, _ = UserDashboardLayout.objects.get_or_create(usuario=request.user)
	hidden_dashboard_card_keys = _get_hidden_dashboard_card_keys(layout)

	items_qs = DashboardItem.objects.filter(ativo=True).prefetch_related("campos")
	items_by_id = {item.id: item for item in items_qs}
	ordered_ids = [item_id for item_id in layout.ordem_itens if item_id in items_by_id]
	remaining_ids = [item.id for item in items_qs if item.id not in ordered_ids]
	ordered_items = [items_by_id[item_id] for item_id in ordered_ids + remaining_ids]
	hidden_item_ids = {item_id for item_id in layout.itens_ocultos if item_id in items_by_id}

	context = {
		"is_admin": is_admin,
		"layout": layout,
		"ordered_items": ordered_items,
		"hidden_item_ids": hidden_item_ids,
		"dashboard_cards": SUMMARY_DASHBOARD_CARDS,
		"hidden_dashboard_card_keys": sorted(hidden_dashboard_card_keys),
		"employees_payloads": get_active_employee_payloads(),
	}
	return render(request, "health/painel/painel.html", context)


@login_required
@require_POST
def importar_funcionarios_planilha(request):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode importar planilha.")

	spreadsheet = request.FILES.get("planilha")
	use_default = request.POST.get("use_default") == "1"
	redirect_target = request.POST.get("next", "")
	if not redirect_target.startswith("/"):
		redirect_target = ""

	if not spreadsheet and not use_default:
		messages.error(request, "Selecione um arquivo de planilha para importar.")
		return redirect(redirect_target or "health:painel")

	try:
		if use_default and not spreadsheet:
			data_dir = Path(settings.BASE_DIR) / "health" / "static" / "health" / "data-js"
			candidates = sorted(data_dir.glob("*.xlsx"))
			if not candidates:
				messages.error(request, "Nenhuma planilha .xlsx encontrada em data-js.")
				return redirect(redirect_target or "health:painel")
			planilha_path = None
			for candidate in candidates:
				if "controle" in candidate.name.lower():
					planilha_path = candidate
					break
			if planilha_path is None:
				planilha_path = candidates[0]
			rows = _read_employee_spreadsheet_path(planilha_path)
		else:
			rows = _read_employee_spreadsheet(spreadsheet)
	except (RuntimeError, ValueError) as exc:
		messages.error(request, str(exc))
		return redirect(redirect_target or "health:painel")
	except Exception as exc:
		messages.error(request, f"Falha ao processar planilha: {exc}")
		return redirect(redirect_target or "health:painel")

	created_count = 0
	updated_count = 0
	skipped_count = 0
	row_errors = []
	next_order = (EmployeeRecord.objects.order_by("-ordem").values_list("ordem", flat=True).first() or 0) + 1
	grouped_payloads = {}
	grouped_order = {}

	def merge_payloads(base_payload, incoming_payload):
		for key in ("chapa", "funcao", "secao"):
			if not base_payload.get(key) and incoming_payload.get(key):
				base_payload[key] = incoming_payload.get(key)

		if base_payload.get("qtd_atestados") is None and incoming_payload.get("qtd_atestados") is not None:
			base_payload["qtd_atestados"] = incoming_payload.get("qtd_atestados")

		base_saude = base_payload.get("saudeOcupacional", {})
		incoming_saude = incoming_payload.get("saudeOcupacional", {})
		base_atestados = list(base_saude.get("atestados") or [])
		incoming_atestados = list(incoming_saude.get("atestados") or [])
		if incoming_atestados:
			base_atestados.extend(incoming_atestados)
			base_saude["atestados"] = base_atestados
			base_saude["absenteismo"] = calculate_absenteeism_from_atestados(base_atestados)
			base_payload["saudeOcupacional"] = base_saude

		return base_payload

	for index, raw_row in enumerate(rows, start=2):
		try:
			payload, row = _build_employee_payload_from_row(raw_row)
		except ValueError as exc:
			row_errors.append(f"Linha {index}: {exc}")
			continue

		if not payload.get("nome"):
			skipped_count += 1
			continue

		record_id = _to_optional_int(row.get("id"))
		ativo = _to_bool(row.get("ativo")) if _safe_text(row.get("ativo")) else True
		ordem = _to_optional_int(row.get("ordem"))

		if record_id:
			updated = EmployeeRecord.objects.filter(id=record_id).update(
				payload=payload,
				ativo=ativo,
				ordem=ordem if ordem is not None else next_order,
			)
			if updated:
				updated_count += 1
				if ordem is None:
					next_order += 1
				continue

		key = _normalize_employee_key(payload.get("nome"))
		if key in grouped_payloads:
			grouped_payloads[key] = merge_payloads(grouped_payloads[key], payload)
			continue

		grouped_payloads[key] = payload
		grouped_order[key] = ordem

	if grouped_payloads:
		for key, payload in grouped_payloads.items():
			ordem = grouped_order.get(key)
			EmployeeRecord.objects.create(
				payload=payload,
				ativo=True,
				ordem=ordem if ordem is not None else next_order,
			)
			created_count += 1
			if ordem is None:
				next_order += 1

	if created_count or updated_count:
		messages.success(
			request,
			f"Importação concluída. Criados: {created_count}, atualizados: {updated_count}, ignorados: {skipped_count}.",
		)
	else:
		messages.warning(request, "Nenhum funcionário foi importado. Verifique as colunas da planilha.")

	if row_errors:
		messages.warning(request, "Ocorreram erros em algumas linhas: " + " | ".join(row_errors[:3]))

	return redirect(redirect_target or "health:painel")


@login_required
def exportar_funcionarios_json(request):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode exportar dados.")

	records = EmployeeRecord.objects.all().order_by("ordem", "id")
	data = []
	for record in records:
		payload = normalize_employee_payload(record.payload if isinstance(record.payload, dict) else {})
		data.append({
			"id": record.id,
			"ativo": record.ativo,
			"ordem": record.ordem,
			"payload": payload,
		})

	content = json.dumps(data, ensure_ascii=False, indent=2)
	response = HttpResponse(content, content_type="application/json; charset=utf-8")
	response["Content-Disposition"] = 'attachment; filename="funcionarios.json"'
	return response


@login_required
def exportar_funcionarios_xml(request):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode exportar dados.")

	records = EmployeeRecord.objects.all().order_by("ordem", "id")
	xml_content = _build_employees_xml(records)
	response = HttpResponse(xml_content, content_type="application/xml; charset=utf-8")
	response["Content-Disposition"] = 'attachment; filename="funcionarios.xml"'
	return response


@login_required
@require_POST
def salvar_layout(request):
	try:
		payload = json.loads(request.body.decode("utf-8"))
	except (TypeError, json.JSONDecodeError):
		return JsonResponse({"ok": False, "error": "Payload inválido."}, status=400)

	order = payload.get("order", [])
	view_mode = payload.get("view_mode", "cards")
	hidden_ids = payload.get("hidden_ids", [])

	if not isinstance(order, list):
		return JsonResponse({"ok": False, "error": "Ordem inválida."}, status=400)

	if not isinstance(hidden_ids, list):
		return JsonResponse({"ok": False, "error": "Lista de ocultos inválida."}, status=400)

	hidden_dashboard_cards = payload.get("hidden_dashboard_cards", [])
	if not isinstance(hidden_dashboard_cards, list):
		return JsonResponse({"ok": False, "error": "Lista de cards ocultos inválida."}, status=400)

	valid_mode = view_mode in {"cards", "grid"}
	if not valid_mode:
		return JsonResponse({"ok": False, "error": "Modo inválido."}, status=400)

	layout, _ = UserDashboardLayout.objects.get_or_create(usuario=request.user)
	layout.ordem_itens = [int(item_id) for item_id in order if str(item_id).isdigit()]
	layout.itens_ocultos = [int(item_id) for item_id in hidden_ids if str(item_id).isdigit()]
	layout.cards_ocultos_dashboard = [str(card_key) for card_key in hidden_dashboard_cards if str(card_key) in SUMMARY_DASHBOARD_CARD_KEYS]
	layout.modo_visualizacao = view_mode
	layout.save(update_fields=["ordem_itens", "modo_visualizacao", "itens_ocultos", "cards_ocultos_dashboard"])

	return JsonResponse({"ok": True})


@login_required
@require_POST
def adicionar_ou_atualizar_info(request, campo_id):
	campo = get_object_or_404(DashboardField, id=campo_id)
	valor = request.POST.get("valor", "").strip()

	if not valor:
		messages.error(request, "Informe um valor antes de salvar.")
		return redirect("health:painel")

	DashboardEntry.objects.update_or_create(
		campo=campo,
		usuario=request.user,
		defaults={"valor": valor},
	)
	messages.success(request, "Informação salva com sucesso.")
	return redirect("health:painel")


@login_required
@require_POST
def editar_info(request, entry_id):
	entry = get_object_or_404(DashboardEntry, id=entry_id)
	is_admin = user_is_admin(request.user)

	if not is_admin and entry.usuario_id != request.user.id:
		return HttpResponseForbidden("Sem permissão para editar esta informação.")

	valor = request.POST.get("valor", "").strip()
	if not valor:
		messages.error(request, "Informe um valor válido para atualizar.")
		return redirect("health:painel")

	entry.valor = valor
	entry.save(update_fields=["valor", "atualizado_em"])
	messages.success(request, "Informação atualizada.")
	return redirect("health:painel")


@login_required
@require_POST
def remover_info(request, entry_id):
	entry = get_object_or_404(DashboardEntry, id=entry_id)
	is_admin = user_is_admin(request.user)

	if not is_admin and entry.usuario_id != request.user.id:
		return HttpResponseForbidden("Sem permissão para remover esta informação.")

	entry.delete()
	messages.success(request, "Informação removida.")
	return redirect("health:painel")


@login_required
@require_POST
def renomear_item(request, item_id):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode renomear itens.")

	item = get_object_or_404(DashboardItem, id=item_id)
	novo_nome = request.POST.get("nome", "").strip()
	if not novo_nome:
		messages.error(request, "Nome do item não pode ficar vazio.")
		return redirect("health:painel")

	item.nome = novo_nome
	item.save(update_fields=["nome"])
	messages.success(request, "Item renomeado.")
	return redirect("health:painel")


@login_required
@require_POST
def renomear_campo(request, campo_id):
	if not user_is_admin(request.user):
		return HttpResponseForbidden("Somente administrador pode renomear campos.")

	campo = get_object_or_404(DashboardField, id=campo_id)
	novo_nome = request.POST.get("nome", "").strip()
	if not novo_nome:
		messages.error(request, "Nome do campo não pode ficar vazio.")
		return redirect("health:painel")

	campo.nome = novo_nome
	campo.save(update_fields=["nome"])
	messages.success(request, "Campo renomeado.")
	return redirect("health:painel")
